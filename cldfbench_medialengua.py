from collections import ChainMap
import pathlib
import sys

from clldutils import sfm
from openpyxl import load_workbook
from pydictionaria.formats import sfm as formats_sfm
from pydictionaria.formats.sfm_lib import Database as SFM
from pydictionaria import sfm2cldf

from cldfbench import CLDFSpec, Dataset as BaseDataset


# Excel->SFM conversion

HEADER_MAPPING = (
    ('headword', 'lx'),
    ('pronuncation', 'ph'),
    ('audio', 'sf'),
    ('part-of-speech', 'ps'),
    ('English', 'de'),
    ('Spanish', 'gxx'),
    ('Quichua', 'gxy'),
    ('source language', 'et'),
    ('synomyms', 'sy'),
    ('variant', 'va'),
    ('general comments', 'nt'),
    ('source', 'src'),
    ('sample sentence from source', 'xv'),
    ('English translation', 'xe'))

LIST_MARKERS = ('sy', 'va')


def _dictionarise(column_names, row):
    return dict(
        (name, str(cell.value))
        for name, cell in zip(column_names, row)
        if name and cell.value)


def _normalise_cell(cell):
    if isinstance(cell, str):
        return re.sub(r'\s+', ' ', cell.strip())
    else:
        return cell


def normalise_whitespace(row):
    """Return table row with normalised white space.

    This involves stripping leading and trailing whitespace, as well as
    consolidating white space to single spaces.
    """
    pairs = (
        (k, _normalise_cell(v))
        for k, v in row.items())
    return {
        k: v for k, v in pairs
        if not isinstance(v, str) or v}


def sheet_to_list(worksheet):
    """Return an Excel worksheet as list of dictionaries.

    Each dictionary in the list represents a row of the worksheet.  The
    dictionaries themselvs map column names to values.
    """
    column_names = [cell.value for cell in worksheet['1']]
    data = worksheet['2:{}'.format(worksheet.max_row)]
    rows = [_dictionarise(column_names, row) for row in data]
    return [row for row in rows if row]


def excel_row_to_sfm(row, header_mapping):
    return sfm.Entry(
        (marker, row[header])
        for header, marker in header_mapping
        if header in row)


def _reorganize_list_field(pair, list_markers):
    marker, value = pair
    if marker in list_markers:
        new_val = '; '.join(v.strip() for v in value.split(',') if v.strip())
        return marker, new_val
    else:
        return pair


def reorganize_lists(entry, list_markers):
    return entry.__class__(
        _reorganize_list_field(pair, list_markers) for pair in entry)


# Preprocessing

def reorganize(sfm):
    return sfm


def preprocess(entry):
    return entry


# Postprocessing

def authors_string(authors):
    def is_primary(a):
        return not isinstance(a, dict) or a.get('primary', True)

    primary = ' and '.join(
        a['name'] if isinstance(a, dict) else a
        for a in authors
        if is_primary(a))
    secondary = ' and '.join(
        a['name']
        for a in authors
        if not is_primary(a))
    if primary and secondary:
        return '{} with {}'.format(primary, secondary)
    else:
        return primary or secondary


class Dataset(BaseDataset):
    dir = pathlib.Path(__file__).parent
    id = "medialengua"

    def cldf_specs(self):  # A dataset must declare all CLDF sets it creates.
        return CLDFSpec(
            dir=self.cldf_dir,
            module='Dictionary',
            metadata_fname='cldf-metadata.json')

    def cmd_download(self, args):
        """
        Download files to the raw/ directory. You can use helpers methods of `self.raw_dir`, e.g.

        >>> self.raw_dir.download(url, fname)
        """
        workbook = load_workbook(filename=self.raw_dir / 'db-original.xlsx')
        table_data = sheet_to_list(workbook[workbook.sheetnames[0]])
        sfm_data = sfm.SFM(
            excel_row_to_sfm(row, HEADER_MAPPING)
            for row in table_data)
        sfm_data.visit(lambda e: reorganize_lists(e, LIST_MARKERS))
        sfm_data.write(self.raw_dir / 'db.sfm')

    def cmd_makecldf(self, args):
        """
        Convert the raw data to a CLDF dataset.

        >>> args.writer.objects['LanguageTable'].append(...)
        """

        # read data

        md = self.etc_dir.read_json('md.json')
        properties = md.get('properties') or {}
        language_name = md['language']['name']
        isocode = md['language']['isocode']
        language_id = md['language']['isocode']
        glottocode = md['language']['glottocode']

        marker_map = ChainMap(
            properties.get('marker_map') or {},
            formats_sfm.DEFAULT_MARKER_MAP)
        entry_sep = properties.get('entry_sep') or sfm2cldf.DEFAULT_ENTRY_SEP
        sfm = SFM(
            self.raw_dir / 'db.sfm',
            marker_map=marker_map,
            entry_sep=entry_sep)

        examples = formats_sfm.load_examples(self.raw_dir / 'examples.sfm')

        if (self.etc_dir / 'cdstar.json').exists():
            media_catalog = self.etc_dir.read_json('cdstar.json')
        else:
            media_catalog = {}

        # preprocessing

        sfm = reorganize(sfm)
        sfm.visit(preprocess)

        # processing

        with open(self.dir / 'cldf.log', 'w', encoding='utf-8') as log_file:
            log_name = '%s.cldf' % language_id
            cldf_log = sfm2cldf.make_log(log_name, log_file)

            entries, senses, examples, media = sfm2cldf.process_dataset(
                self.id, language_id, properties,
                sfm, examples, media_catalog=media_catalog,
                glosses_path=self.raw_dir / 'glosses.flextext',
                examples_log_path=self.dir / 'examples.log',
                glosses_log_path=self.dir / 'glosses.log',
                cldf_log=cldf_log)

            # good place for some post-processing

            # cldf schema

            sfm2cldf.make_cldf_schema(
                args.writer.cldf, properties,
                entries, senses, examples, media)

            sfm2cldf.attach_column_titles(args.writer.cldf, properties)

            print(file=log_file)

            entries = sfm2cldf.ensure_required_columns(
                args.writer.cldf, 'EntryTable', entries, cldf_log)
            senses = sfm2cldf.ensure_required_columns(
                args.writer.cldf, 'SenseTable', senses, cldf_log)
            examples = sfm2cldf.ensure_required_columns(
                args.writer.cldf, 'ExampleTable', examples, cldf_log)
            media = sfm2cldf.ensure_required_columns(
                args.writer.cldf, 'media.csv', media, cldf_log)

            entries = sfm2cldf.remove_senseless_entries(
                senses, entries, cldf_log)

        # output

        args.writer.cldf.properties['dc:creator'] = authors_string(
            md.get('authors') or ())

        language = {
            'ID': language_id,
            'Name': language_name,
            'ISO639P3code': isocode,
            'Glottocode': glottocode,
        }
        args.writer.objects['LanguageTable'] = [language]

        args.writer.objects['EntryTable'] = entries
        args.writer.objects['SenseTable'] = senses
        args.writer.objects['ExampleTable'] = examples
        args.writer.objects['media.csv'] = media
